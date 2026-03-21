import json

import pandas as pd
from openpyxl.styles import PatternFill

ITEMS_BASE_PATH = r"..\data\items-base.json"
MAGIC_VARIANTS_PATH = r"..\data\magicvariants.json"
ITEMS_PATH = r"..\data\items.json"
OUTPUT_FILE = r"..\..\Excel\items.xlsx"


def _jsonify(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _entries_to_text(entries):
    if isinstance(entries, list):
        return "\n".join(entry for entry in entries if isinstance(entry, str)) or None
    if isinstance(entries, str):
        return entries
    return None


def _summarize_requires(requires):
    if not isinstance(requires, list):
        return None
    out = []
    for req in requires:
        if not isinstance(req, dict):
            continue
        if "type" in req:
            out.append(f"type={req['type']}")
        for key, value in req.items():
            if isinstance(value, bool) and value:
                out.append(f"{key}=true")
    return "; ".join(dict.fromkeys(out)) or None


def _load_base_items():
    with open(ITEMS_BASE_PATH, "r", encoding="utf-8") as f_in:
        raw = json.load(f_in)
    base_items = [item for item in raw.get("baseitem", []) if item]

    for item in base_items:
        item["entries_text"] = _entries_to_text(item.get("entries"))

    data = pd.DataFrame(base_items).map(_jsonify)
    data.insert(1, "Allowed_By_Nuno", False)
    data.insert(2, "Iconify_id", None)
    return data


def _load_variant_items():
    with open(MAGIC_VARIANTS_PATH, "r", encoding="utf-8") as f_in:
        raw = json.load(f_in)
    variants = [variant for variant in raw.get("magicvariant", []) if variant]

    variant_rows = []
    for variant in variants:
        inherits = variant.get("inherits", {})
        row = {
            "name": variant.get("name"),
            "type": variant.get("type"),
            "requires_summary": _summarize_requires(variant.get("requires")),
            "requires": variant.get("requires"),
            "excludes": variant.get("excludes"),
            "entries_text": _entries_to_text(variant.get("entries")),
            "namePrefix": inherits.get("namePrefix"),
            "source": inherits.get("source"),
            "page": inherits.get("page"),
            "tier": inherits.get("tier"),
            "rarity": inherits.get("rarity"),
            "bonusAc": inherits.get("bonusAc"),
            "bonusWeapon": inherits.get("bonusWeapon"),
            "bonusWeaponAttack": inherits.get("bonusWeaponAttack"),
            "inherits_entries_text": _entries_to_text(inherits.get("entries")),
            "inherits": inherits,
        }
        variant_rows.append(row)

    data = pd.DataFrame(variant_rows).map(_jsonify)
    data.insert(1, "Allowed_By_Nuno", False)
    data.insert(2, "Iconify_id", None)
    return data


def _load_items():
    with open(ITEMS_PATH, "r", encoding="utf-8") as f_in:
        raw = json.load(f_in)
    items = [item for item in raw.get("item", []) if item]

    for item in items:
        item["entries_text"] = _entries_to_text(item.get("entries"))

    data = pd.DataFrame(items).map(_jsonify)
    data.insert(1, "Allowed_By_Nuno", False)
    data.insert(2, "Iconify_id", None)
    return data


def _apply_checkbox_format(worksheet, data_rows, red_fill, green_fill):
    for row in range(2, data_rows + 2):
        cell = worksheet.cell(row=row, column=2)
        if cell.value is True:
            cell.value = "✓"
            cell.fill = green_fill
        elif cell.value is False:
            cell.value = "✗"
            cell.fill = red_fill


def main():
    base_items_df = _load_base_items()
    variants_df = _load_variant_items()
    items_df = _load_items()

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")

        base_sheet = "base_items"
        variant_sheet = "variants"
        items_sheet = "items"

        base_items_df.to_excel(writer, sheet_name=base_sheet, index=False)
        variants_df.to_excel(writer, sheet_name=variant_sheet, index=False)
        items_df.to_excel(writer, sheet_name=items_sheet, index=False)

        _apply_checkbox_format(writer.sheets[base_sheet], len(base_items_df), red_fill, green_fill)
        _apply_checkbox_format(writer.sheets[variant_sheet], len(variants_df), red_fill, green_fill)
        _apply_checkbox_format(writer.sheets[items_sheet], len(items_df), red_fill, green_fill)

        print(f"Wrote {OUTPUT_FILE} with sheets '{base_sheet}', '{variant_sheet}', and '{items_sheet}'")


if __name__ == "__main__":
    main()
