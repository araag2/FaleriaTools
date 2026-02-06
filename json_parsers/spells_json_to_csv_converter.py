import json
import pandas as pd
import os
from tqdm import tqdm
from openpyxl.styles import PatternFill

ALLOWED_FILES = [
    "spells-phb.json",
    "spells-egw.json",
    "spells-ftd.json",
    "spells-ggr.json",
    "spells-idrotf.json",
    "spells-llk.json",
    "spells-sato.json",
    "spells-scc.json",
    "spells-tce.json",
    "spells-xge.json",
    "spells-bmt.json",
    "spells-aag.json"
]

COLUMNS_TO_DELETE = [
    "time",
    "range",
    "duration",
    "meta"
]

COLUMNS_TO_MOVE_TO_END = [
    "components",
    "entries",
    "scalingLevelDice",
    "damageInflict",
    "savingThrow",
    "miscTags",
    "areaTags"
]

def split_time_field(spell_entry):
    spell_entry["time_number"] = None
    spell_entry["time_unit"] = None
    spell_entry["time_condition"] = None

    time_entry = spell_entry.get("time", None)[0]

    if time_entry:
        spell_entry["time_number"] = time_entry["number"] if "number" in time_entry else None
        spell_entry["time_unit"] = time_entry["unit"] if "unit" in time_entry else None
        spell_entry["time_condition"] = time_entry["condition"] if "condition" in time_entry else None

    return spell_entry 

def split_range_field(spell_entry):
    spell_entry["range_type"] = None
    spell_entry["range_distance_self"] = False
    spell_entry["range_distance_feet"] = None
    spell_entry["range_distance_meters"] = None
    spell_entry["range_distance_squares"] = None

    range_entry = spell_entry.get("range", None)
    range_distance = range_entry.get("distance", None) if range_entry else None

    if range_entry:
        spell_entry["range_type"] = range_entry["type"] if "type" in range_entry else None

    if range_distance:
        spell_entry["range_distance_self"] = True if "type" in range_distance and range_distance["type"] == "self" else False

        spell_entry["range_distance_feet"] = range_distance["amount"] if "amount" in range_distance else None

        if spell_entry["range_distance_feet"] is not None:
            spell_entry["range_distance_meters"] = round(spell_entry["range_distance_feet"] * 0.3, 1)
            spell_entry["range_distance_squares"] = round(spell_entry["range_distance_feet"] / 5)

    return spell_entry

def split_components_field(spell_entry):
    spell_entry["components_verbal"] = False
    spell_entry["components_somatic"] = False
    spell_entry["components_material"] = False
    spell_entry["components_material_description"] = None
    spell_entry["components_material_gc_cost"] = None
    spell_entry["components_material_consume"] = None

    components_entry = spell_entry.get("components", None)

    if components_entry:
        spell_entry["components_verbal"] = True if "v" in components_entry else False
        spell_entry["components_somatic"] = True if "s" in components_entry else False

        if "m" in components_entry:
            spell_entry["components_material"] = True

            if isinstance(components_entry["m"], str):
                spell_entry["components_material_description"] = components_entry["m"]
            elif isinstance(components_entry["m"], dict):
                spell_entry["components_material_description"] = components_entry["m"].get("text", None)
                spell_entry["components_material_gc_cost"] = components_entry["m"].get("cost", None)
                spell_entry["components_material_consume"] = components_entry["m"].get("consume", False)

    return spell_entry    

def split_duration_field(spell_entry):
    spell_entry["duration_type"] = None
    spell_entry["duration_concentration"] = False
    spell_entry["duration_time_number"] = None
    spell_entry["duration_time_unit"] = None
    spell_entry["duration_end_condition"] = None

    duration_entry = spell_entry.get("duration", None)
    if duration_entry and isinstance(duration_entry, list) and duration_entry:
        duration_entry = duration_entry[0]

        if duration_entry:
            spell_entry["duration_type"] = duration_entry.get("type", None)
            spell_entry["duration_concentration"] = duration_entry.get("concentration", False)
            spell_entry["duration_end_condition"] = duration_entry.get("ends", None)

            if "duration" in duration_entry:
                time_entry = duration_entry["duration"]
                spell_entry["duration_time_number"] = time_entry.get("amount", None)
                spell_entry["duration_time_unit"] = time_entry.get("type", None)

    return spell_entry

def split_entries_text_field(spell_entry):
    spell_entry["entries_text"] = None

    entries = spell_entry.get("entries", None)

    if entries:
        spell_entry["entries_text"] = "\n".join([entry for entry in entries if isinstance(entry, str)]) if isinstance(entries, list) else entries

    return spell_entry

def split_entries_higher_level_field(spell_entry):
    spell_entry["entriesHigherLevel"] = None

    entries_higher_level = spell_entry.get("entriesHigherLevel", None)

    if entries_higher_level:
        spell_entry["entriesHigherLevel"] = "\n".join(entries_higher_level[0]["entries"]) if "entries" in entries_higher_level[0] else None

    return spell_entry

def split_scaling_level_dice_field(spell_entry):
    spell_entry["scalingLevelDice_dmg_type"] = None
    spell_entry["scalingLevelDice_dice_lvl1"] = None
    spell_entry["scalingLevelDice_dice_lvl5"] = None
    spell_entry["scalingLevelDice_dice_lvl11"] = None
    spell_entry["scalingLevelDice_dice_lvl17"] = None

    scaling_entry = spell_entry.get("scalingLevelDice", None)

    if scaling_entry:
        spell_entry["scalingLevelDice_dmg_type"] = scaling_entry["label"] if "label" in scaling_entry else None

        if "scaling" in scaling_entry:
            spell_entry["scalingLevelDice_dice_lvl1"] = scaling_entry["scaling"].get("1", None)
            spell_entry["scalingLevelDice_dice_lvl5"] = scaling_entry["scaling"].get("5", None)
            spell_entry["scalingLevelDice_dice_lvl11"] = scaling_entry["scaling"].get("11", None)
            spell_entry["scalingLevelDice_dice_lvl17"] = scaling_entry["scaling"].get("17", None)

    return spell_entry

def split_meta_ritual_field(spell_entry):
    spell_entry["ritual_cast"] = False

    meta_entry = spell_entry.get("meta", None)

    if meta_entry and "ritual" in meta_entry:
        spell_entry["ritual_cast"] = True if meta_entry["ritual"] == True else False

    return spell_entry

FIELD_FUNCTIONS = [
    split_time_field,
    split_range_field,
    split_components_field,
    split_duration_field,
    split_entries_text_field,
    split_entries_higher_level_field,
    split_scaling_level_dice_field,
    split_meta_ritual_field
]

def process_spell_file(file_path):
    with open(file_path, "r", encoding='utf-8') as f_in:
        raw = json.load(f_in)
        if "spell" not in raw:
            return None
        spells = raw["spell"]
        
        # Filter out empty spell entries
        spells = [spell for spell in spells if spell]
        
        for spell in tqdm(spells, desc=f"Processing {os.path.basename(file_path)}"):
            for func in FIELD_FUNCTIONS:
                spell = func(spell)

        data = pd.DataFrame(spells)

        # Delete unwanted columns
        data.drop(columns=COLUMNS_TO_DELETE, inplace=True, errors='ignore')

        # Add a 2nd column
        data.insert(1, "Allowed_By_Nuno", False)
        data.insert(2, "Iconify_id", None)

        return data

def main():
    directory = r"..\data\spells"
    output_file = r"..\..\Excel\spells.xlsx"

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
        
        for root, dirs, files in os.walk(directory):
            for file in files:
                if file.endswith('.json') and file in ALLOWED_FILES:
                    file_path = os.path.join(root, file)
                    sheet_name = os.path.splitext(file)[0]
                    data = process_spell_file(file_path)
                    if data is not None:
                        data.to_excel(writer, sheet_name=sheet_name, index=False)
                        worksheet = writer.sheets[sheet_name]
                        
                        # Apply tickbox formatting to column B (Allowed_By_Nuno)
                        for row in range(2, len(data) + 2):  # Data starts at row 2 (1-indexed)
                            cell = worksheet.cell(row=row, column=2)
                            if cell.value is True:
                                cell.value = "✓"
                                cell.fill = green_fill
                            elif cell.value is False:
                                cell.value = "✗"
                                cell.fill = red_fill
                        print(f"Processed {file_path} into sheet '{sheet_name}'")

if __name__ == "__main__":
    main()