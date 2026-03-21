import json
import os

import pandas as pd
from openpyxl.styles import PatternFill
from tqdm import tqdm

ALLOWED_FILES = [
    "feats.json",
]


def _jsonify(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _extract_entries_text(feat_entry):
    feat_entry["entries_text"] = None
    entries = feat_entry.get("entries")
    if isinstance(entries, list):
        feat_entry["entries_text"] = "\n".join(entry for entry in entries if isinstance(entry, str)) or None
    return feat_entry


def _extract_prerequisite_text(feat_entry):
    feat_entry["prerequisite_text"] = None
    prerequisite = feat_entry.get("prerequisite")
    if prerequisite:
        feat_entry["prerequisite_text"] = json.dumps(prerequisite, ensure_ascii=False)
    return feat_entry


def _extract_ability_text(feat_entry):
    feat_entry["ability_text"] = None
    ability = feat_entry.get("ability")
    if ability:
        feat_entry["ability_text"] = json.dumps(ability, ensure_ascii=False)
    return feat_entry


FIELD_FUNCTIONS = [
    _extract_entries_text,
    _extract_prerequisite_text,
    _extract_ability_text,
]


def process_feat_file(file_path):
    with open(file_path, "r", encoding="utf-8") as f_in:
        raw = json.load(f_in)
        if "feat" not in raw:
            return None

        feats = [feat for feat in raw["feat"] if feat]
        for feat in tqdm(feats, desc=f"Processing {os.path.basename(file_path)}"):
            for func in FIELD_FUNCTIONS:
                feat = func(feat)

        data = pd.DataFrame(feats)
        data = data.map(_jsonify)
        data.insert(1, "Allowed_By_Nuno", False)
        data.insert(2, "Iconify_id", None)
        return data


def main():
    directory = r"..\data"
    output_file = r"..\..\Excel\feats.xlsx"

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")

        for root, _, files in os.walk(directory):
            for file in files:
                if file.endswith(".json") and file in ALLOWED_FILES:
                    file_path = os.path.join(root, file)
                    data = process_feat_file(file_path)
                    if data is None:
                        continue

                    sheet_name = "feats"
                    data.to_excel(writer, sheet_name=sheet_name, index=False)
                    worksheet = writer.sheets[sheet_name]

                    for row in range(2, len(data) + 2):
                        cell = worksheet.cell(row=row, column=2)
                        if cell.value is True:
                            cell.value = "✓"
                            cell.fill = green_fill
                        elif cell.value is False:
                            cell.value = "✗"
                            cell.fill = red_fill

                    print(f"Processed {file_path} into sheet '{sheet_name}'")


if __name__ == "__main__":
    main()
